from __future__ import annotations

import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets

import resend
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen
from uuid import uuid4

from flask import Response, abort, flash, redirect, render_template_string, request, session, url_for
from sqlalchemy import Index

_models = {}


def init_feature_models(db):
    if _models:
        return _models

    class StudentReportPayment(db.Model):
        __tablename__ = "student_report_payments"
        id = db.Column(db.Integer, primary_key=True)
        school_id = db.Column(db.Integer, db.ForeignKey("schools.id", ondelete="CASCADE"), nullable=False, index=True)
        user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
        student_id = db.Column(db.Integer, db.ForeignKey("students.id", ondelete="CASCADE"), nullable=False, index=True)
        academic_year = db.Column(db.String(40), nullable=False, index=True)
        term = db.Column(db.String(40), nullable=False, index=True)
        reference = db.Column(db.String(120), nullable=False, unique=True, index=True)
        amount_subunit = db.Column(db.Integer, nullable=False)
        currency = db.Column(db.String(10), nullable=False)
        status = db.Column(db.String(30), default="pending", nullable=False, index=True)
        authorization_url = db.Column(db.String(500), default="")
        provider_transaction_id = db.Column(db.String(80), default="")
        paid_at = db.Column(db.DateTime)
        created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
        updated_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    class PasswordResetToken(db.Model):
        __tablename__ = "password_reset_tokens"
        id = db.Column(db.Integer, primary_key=True)
        user_id = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="CASCADE"), nullable=False, index=True)
        token_hash = db.Column(db.String(64), nullable=False, unique=True, index=True)
        expires_at = db.Column(db.DateTime, nullable=False, index=True)
        used_at = db.Column(db.DateTime)
        requested_ip = db.Column(db.String(80), default="")
        created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    class BulkImportJob(db.Model):
        __tablename__ = "bulk_import_jobs"
        id = db.Column(db.Integer, primary_key=True)
        school_id = db.Column(db.Integer, db.ForeignKey("schools.id", ondelete="CASCADE"), nullable=False, index=True)
        created_by = db.Column(db.Integer, db.ForeignKey("users.id", ondelete="SET NULL"), index=True)
        record_type = db.Column(db.String(20), nullable=False)
        filename = db.Column(db.String(260), nullable=False)
        status = db.Column(db.String(30), default="preview", nullable=False, index=True)
        duplicate_mode = db.Column(db.String(20), default="skip", nullable=False)
        staged_json = db.Column(db.Text, nullable=False)
        report_json = db.Column(db.Text, default="{}")
        created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
        completed_at = db.Column(db.DateTime)

    Index("ix_student_report_unlock", StudentReportPayment.school_id, StudentReportPayment.user_id,
          StudentReportPayment.student_id, StudentReportPayment.academic_year,
          StudentReportPayment.term, StudentReportPayment.status)
    _models.update(StudentReportPayment=StudentReportPayment,
                   PasswordResetToken=PasswordResetToken,
                   BulkImportJob=BulkImportJob)
    return _models


def _subunit(value):
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, TypeError) as exc:
        raise RuntimeError("The configured student report fee is invalid") from exc
    if amount < 0:
        raise RuntimeError("The student report fee cannot be negative")
    return int((amount * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _paystack(secret, path, method="GET", payload=None):
    if not secret:
        raise RuntimeError("Paystack is not configured")
    req = Request("https://api.paystack.co" + path,
                  data=json.dumps(payload).encode() if payload is not None else None,
                  method=method,
                  headers={"Authorization": "Bearer " + secret, "Content-Type": "application/json",
                           "Accept": "application/json"})
    try:
        with urlopen(req, timeout=20) as response:
            result = json.loads(response.read().decode())
    except (HTTPError, URLError, TimeoutError, ValueError) as exc:
        raise RuntimeError("The payment service could not be reached") from exc
    if not result.get("status"):
        raise RuntimeError(result.get("message") or "Paystack rejected the request")
    return result


def _clean_env(name):
    """Read Render variables without accidental whitespace or wrapping quotes."""
    return os.getenv(name, "").strip().strip('"').strip("'")


def _send_resend(to_email, reset_url, user_name):
    api_key = _clean_env("RESEND_API_KEY")
    sender = _clean_env("EMAIL_FROM")
    recipient = (to_email or "").strip().lower()
    if not api_key or not sender:
        raise RuntimeError("RESEND_API_KEY and EMAIL_FROM must be configured")
    if not recipient:
        raise RuntimeError("The administrator account has no email address")

    resend.api_key = api_key
    params = {
        "from": sender,
        "to": [recipient],
        "subject": "Reset your Smart Schools SMS password",
        "html": (
            "<p>Hello " + user_name + ",</p>"
            "<p>Use the secure link below to choose a new password. The link expires soon and works once.</p>"
            '<p><a href="' + reset_url + '">Reset my password</a></p>'
            "<p>If you did not request this, you can ignore this email.</p>"
        ),
    }
    try:
        result = resend.Emails.send(params)
    except Exception as exc:
        # Resend's SDK includes the HTTP status and API message (for example,
        # an unverified sender domain) without exposing the API key.
        raise RuntimeError(f"Resend rejected the reset email: {exc}") from exc

    email_id = result.get("id") if isinstance(result, dict) else getattr(result, "id", None)
    if not email_id:
        raise RuntimeError("Resend accepted no email ID for the reset request")


def _rows_from_upload(upload):
    name = (upload.filename or "").lower()
    raw = upload.read()
    if not raw:
        raise ValueError("The uploaded file is empty.")
    if len(raw) > 5 * 1024 * 1024:
        raise ValueError("The import file must be 5 MB or smaller.")
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel support is not installed.") from exc
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        return [dict(zip(headers, row)) for row in values[1:]]
    raise ValueError("Upload a .csv or .xlsx file.")


def _clean_row(row):
    return {str(key or "").strip().lower().replace(" ", "_"): str(value or "").strip()
            for key, value in row.items()}


def _validate_rows(rows, record_type, classes):
    required = {"student": ("full_name", "username", "admission_no", "class"),
                "teacher": ("full_name", "username")}[record_type]
    valid, errors = [], []
    seen_usernames, seen_admissions = set(), set()
    class_names = {item.name.strip().lower(): item.id for item in classes}
    for number, raw in enumerate(rows, start=2):
        row = _clean_row(raw)
        messages = [f"missing {field}" for field in required if not row.get(field)]
        username = row.get("username", "").lower()
        if username in seen_usernames:
            messages.append("duplicate username inside file")
        seen_usernames.add(username)
        if record_type == "student":
            admission = row.get("admission_no", "").upper()
            if admission in seen_admissions:
                messages.append("duplicate admission number inside file")
            seen_admissions.add(admission)
            if row.get("class", "").lower() not in class_names:
                messages.append("class does not exist")
            row["admission_no"] = admission
        email = row.get("email") or row.get("guardian_email")
        if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            messages.append("invalid email address")
        if messages:
            errors.append({"row": number, "errors": messages, "data": row})
        else:
            row["_row"] = number
            valid.append(row)
    return valid, errors


def register_feature_routes(app, ctx):
    db = ctx["db"]
    User, School, Student, ClassRoom = ctx["User"], ctx["School"], ctx["Student"], ctx["ClassRoom"]
    ParentReportPayment = ctx["ParentReportPayment"]
    current_user, current_school = ctx["current_user"], ctx["current_school"]
    login_required, school_required = ctx["login_required"], ctx["school_required"]
    csrf_token, log_action = ctx["csrf_token"], ctx["log_action"]
    generate_password_hash = ctx["generate_password_hash"]
    generate_temporary_password = ctx["generate_temporary_password"]
    models = init_feature_models(db)
    StudentReportPayment = models["StudentReportPayment"]
    PasswordResetToken = models["PasswordResetToken"]
    BulkImportJob = models["BulkImportJob"]

    def render_page(body, **values):
        base = ctx["BASE_HTML"]
        return render_template_string(base.replace("{% block body %}{% endblock %}", body), **values)

    def student_payment(student, school, user):
        return StudentReportPayment.query.filter_by(
            school_id=school.id, user_id=user.id, student_id=student.id,
            academic_year=school.academic_year, term=school.term,
            amount_subunit=_subunit(os.getenv("STUDENT_REPORT_FEE", "10.00")),
            currency=os.getenv("PAYSTACK_CURRENCY", "GHS").upper(), status="success").first()

    def verify_student_payment(payment):
        data = (_paystack(os.getenv("PAYSTACK_SECRET_KEY", ""),
                          "/transaction/verify/" + quote(payment.reference, safe="")).get("data") or {})
        valid = (data.get("status") == "success" and str(data.get("reference")) == payment.reference
                 and int(data.get("amount") or -1) == payment.amount_subunit
                 and str(data.get("currency") or "").upper() == payment.currency)
        payment.status = "success" if valid else "failed"
        payment.provider_transaction_id = str(data.get("id") or "")[:80]
        payment.updated_at = datetime.utcnow()
        if valid and not payment.paid_at:
            payment.paid_at = datetime.utcnow()
        db.session.commit()
        return valid

    ctx["student_report_is_paid"] = student_payment

    @app.route("/forgot-password", methods=["GET", "POST"])
    def forgot_password():
        portal = request.args.get("portal", "admin")
        if request.method == "POST":
            email = request.form.get("email", "").strip().lower()
            admin_roles = {"system_admin", "school_admin", "accountant", "registrar", "librarian", "receptionist"}
            user = User.query.filter(User.email.ilike(email), User.role.in_(admin_roles), User.active.is_(True)).first()
            if user:
                PasswordResetToken.query.filter_by(user_id=user.id, used_at=None).update(
                    {"used_at": datetime.utcnow()}, synchronize_session=False)
                token = secrets.token_urlsafe(32)
                digest = hashlib.sha256(token.encode()).hexdigest()
                minutes = max(5, int(os.getenv("PASSWORD_RESET_EXPIRY_MINUTES", "60")))
                db.session.add(PasswordResetToken(
                    user_id=user.id, token_hash=digest,
                    expires_at=datetime.utcnow() + timedelta(minutes=minutes),
                    requested_ip=request.headers.get("X-Forwarded-For", request.remote_addr or "")))
                db.session.commit()
                base_url = os.getenv("APP_URL", "").rstrip("/")
                reset_url = (base_url + url_for("reset_password_with_token", token=token)
                             if base_url else url_for("reset_password_with_token", token=token, _external=True))
                try:
                    _send_resend(user.email, reset_url, user.full_name)
                except RuntimeError:
                    app.logger.exception("Password reset email delivery failed")
            flash("If that administrator email exists, a reset link has been sent.", "success")
        body = """<main class="login-shell"><section class="card login-card"><h2>Reset administrator password</h2><p class="muted">Enter the email saved on your administrator account.</p>{% for category,message in get_flashed_messages(with_categories=true) %}<div class="flash {{ category }}">{{ message }}</div>{% endfor %}<form method="post">{{ csrf() }}{{ field('Administrator Email','email','email',required=true) }}<button class="btn green">Send reset link</button></form><p><a class="btn ghost" href="{{ url_for('login',portal=portal) }}">Back to login</a></p></section></main>"""
        return render_page(body, title="Forgot Password", portal=portal)

    @app.route("/reset-password/<token>", methods=["GET", "POST"])
    def reset_password_with_token(token):
        digest = hashlib.sha256(token.encode()).hexdigest()
        item = PasswordResetToken.query.filter_by(token_hash=digest, used_at=None).first()
        if not item or item.expires_at < datetime.utcnow():
            return render_page('<main class="login-shell"><section class="card login-card"><h2>Reset link expired</h2><p>Request a new password reset link.</p><a class="btn" href="{{ url_for(\'forgot_password\') }}">Request another link</a></section></main>', title="Expired Link"), 400
        if request.method == "POST":
            password = request.form.get("password", "")
            confirmation = request.form.get("confirmation", "")
            if len(password) < 8:
                flash("Use at least 8 characters.", "error")
            elif password != confirmation:
                flash("The passwords do not match.", "error")
            else:
                user = db.session.get(User, item.user_id)
                user.password_hash = generate_password_hash(password)
                user.must_change_password = False
                item.used_at = datetime.utcnow()
                db.session.commit()
                session.clear()
                flash("Your password has been changed. You can now sign in.", "success")
                return redirect(url_for("login", portal="admin"))
        body = """<main class="login-shell"><section class="card login-card"><h2>Choose a new password</h2>{% for category,message in get_flashed_messages(with_categories=true) %}<div class="flash {{ category }}">{{ message }}</div>{% endfor %}<form method="post">{{ csrf() }}{{ field('New Password','password','password',required=true) }}{{ field('Confirm Password','confirmation','password',required=true) }}<button class="btn green">Change password</button></form></section></main>"""
        return render_page(body, title="Choose New Password")

    @app.route("/student/report/payment")
    @login_required("student")
    @school_required
    def student_report_payment():
        user, school = current_user(), current_school()
        student = Student.query.filter_by(user_id=user.id, school_id=school.id).first_or_404()
        if student_payment(student, school, user):
            return redirect(url_for("student_results"))
        fee = Decimal(_subunit(os.getenv("STUDENT_REPORT_FEE", "10.00"))) / 100
        body = """<main class="wrap"><div class="layout">""" + ctx["SIDEBAR"] + """<section class="card" style="max-width:620px"><h2>Unlock your report</h2><p>{{ school.academic_year }} · {{ school.term }}</p><h3>{{ currency }} {{ '%.2f'|format(fee) }}</h3><form method="post" action="{{ url_for('initialize_student_report_payment') }}">{{ csrf() }}<button class="btn green">Pay securely with Paystack</button></form><p class="muted">Payment is verified on the server. Smart Schools SMS never stores card or Mobile Money details.</p></section></div></main>"""
        return render_page(body, title="Student Report Payment", fee=float(fee),
                           currency=os.getenv("PAYSTACK_CURRENCY", "GHS").upper())

    @app.route("/student/report/paystack", methods=["POST"])
    @login_required("student")
    @school_required
    def initialize_student_report_payment():
        user, school = current_user(), current_school()
        student = Student.query.filter_by(user_id=user.id, school_id=school.id).first_or_404()
        if student_payment(student, school, user):
            return redirect(url_for("student_results"))
        email = user.email or student.guardian_email
        if not email:
            flash("An email address is required for Paystack. Ask your school administrator to add one.", "error")
            return redirect(url_for("student_report_payment"))
        amount = _subunit(os.getenv("STUDENT_REPORT_FEE", "10.00"))
        reference = f"STU-RPT-{school.id}-{user.id}-{student.id}-{uuid4().hex}"
        payment = StudentReportPayment(
            school_id=school.id, user_id=user.id, student_id=student.id,
            academic_year=school.academic_year, term=school.term, reference=reference,
            amount_subunit=amount, currency=os.getenv("PAYSTACK_CURRENCY", "GHS").upper())
        db.session.add(payment)
        db.session.commit()
        try:
            result = _paystack(os.getenv("PAYSTACK_SECRET_KEY", ""), "/transaction/initialize", "POST", {
                "email": email, "amount": str(amount), "currency": payment.currency,
                "reference": reference, "callback_url": url_for("student_paystack_callback", _external=True),
                "channels": ["mobile_money", "card"],
                "metadata": {"kind": "student_report", "payment_id": payment.id,
                             "school_id": school.id, "student_id": student.id,
                             "academic_year": school.academic_year, "term": school.term}})
            checkout = (result.get("data") or {}).get("authorization_url", "")
            parsed = urlparse(checkout)
            if parsed.scheme != "https" or not parsed.hostname or not parsed.hostname.endswith("paystack.com"):
                raise RuntimeError("Invalid Paystack checkout URL")
            payment.authorization_url = checkout
            payment.updated_at = datetime.utcnow()
            db.session.commit()
            return redirect(checkout)
        except RuntimeError:
            payment.status = "failed"
            payment.updated_at = datetime.utcnow()
            db.session.commit()
            flash("Payment could not be started. Please try again.", "error")
            return redirect(url_for("student_report_payment"))

    @app.route("/payments/paystack/student-callback")
    def student_paystack_callback():
        payment = StudentReportPayment.query.filter_by(
            reference=request.args.get("reference", "").strip()).first_or_404()
        try:
            verified = verify_student_payment(payment)
        except RuntimeError:
            verified = False
        user = current_user()
        if verified and user and user.id == payment.user_id:
            flash("Payment verified. Your report is now unlocked.", "success")
            return redirect(url_for("student_results"))
        flash("Payment was not completed or could not be verified.", "error")
        return redirect(url_for("login", portal="student"))

    @app.route("/payments/paystack/unified-webhook", methods=["POST"])
    def paystack_unified_webhook():
        secret = os.getenv("PAYSTACK_SECRET_KEY", "")
        if not secret:
            return Response(status=503)
        raw = request.get_data(cache=False)
        supplied = request.headers.get("x-paystack-signature", "")
        expected = hmac.new(secret.encode(), raw, hashlib.sha512).hexdigest()
        if not hmac.compare_digest(expected, supplied):
            return Response(status=401)
        try:
            event = json.loads(raw.decode())
        except (UnicodeDecodeError, ValueError):
            return Response(status=400)
        if event.get("event") == "charge.success":
            reference = str((event.get("data") or {}).get("reference") or "")
            student_item = StudentReportPayment.query.filter_by(reference=reference).first()
            parent_item = ParentReportPayment.query.filter_by(reference=reference).first()
            try:
                if student_item and student_item.status != "success":
                    verify_student_payment(student_item)
                elif parent_item and parent_item.status != "success":
                    ctx["verify_report_payment"](parent_item)
            except RuntimeError:
                return Response(status=503)
        return Response(status=200)

    @app.route("/admin/import", methods=["GET", "POST"])
    @login_required("school_admin")
    @school_required
    def bulk_import():
        user, school = current_user(), current_school()
        if request.method == "POST":
            record_type = request.form.get("record_type", "student")
            if record_type not in {"student", "teacher"}:
                abort(400)
            upload = request.files.get("file")
            try:
                rows = _rows_from_upload(upload)
                valid, errors = _validate_rows(rows, record_type,
                                                ClassRoom.query.filter_by(school_id=school.id).all())
                job = BulkImportJob(school_id=school.id, created_by=user.id, record_type=record_type,
                                    filename=(upload.filename or "import")[:260],
                                    staged_json=json.dumps({"valid": valid, "errors": errors}))
                db.session.add(job)
                db.session.commit()
                return redirect(url_for("bulk_import_preview", job_id=job.id))
            except ValueError as exc:
                flash(str(exc), "error")
        recent = BulkImportJob.query.filter_by(school_id=school.id).order_by(BulkImportJob.created_at.desc()).limit(10).all()
        body = """<main class="wrap"><div class="layout">""" + ctx["SIDEBAR"] + """<section class="grid"><article class="card"><h2>Import old school records</h2><p>Upload CSV or Excel (.xlsx), preview validation results, then choose how duplicates are handled.</p>{% for category,message in get_flashed_messages(with_categories=true) %}<div class="flash {{ category }}">{{ message }}</div>{% endfor %}<form method="post" enctype="multipart/form-data">{{ csrf() }}<label>Record type<select name="record_type"><option value="student">Students</option><option value="teacher">Teachers</option></select></label><label>File<input type="file" name="file" accept=".csv,.xlsx" required></label><button class="btn green">Validate and preview</button></form><p class="muted">Student columns: full_name, username, admission_no, class, email, phone, guardian_name, guardian_email, guardian_phone, password. Teacher columns: full_name, username, email, phone, password.</p></article><article class="card"><h3>Recent imports</h3><table><tr><th>File</th><th>Type</th><th>Status</th><th>Report</th></tr>{% for item in recent %}<tr><td>{{ item.filename }}</td><td>{{ item.record_type }}</td><td>{{ item.status }}</td><td><a href="{{ url_for('bulk_import_preview',job_id=item.id) }}">Open</a></td></tr>{% endfor %}</table></article></section></div></main>"""
        return render_page(body, title="Bulk Import", recent=recent)

    @app.route("/admin/import/<int:job_id>", methods=["GET", "POST"])
    @login_required("school_admin")
    @school_required
    def bulk_import_preview(job_id):
        user, school = current_user(), current_school()
        job = BulkImportJob.query.filter_by(id=job_id, school_id=school.id).first_or_404()
        staged = json.loads(job.staged_json)
        if request.method == "POST" and job.status == "preview":
            mode = request.form.get("duplicate_mode", "skip")
            if mode not in {"skip", "update"}:
                abort(400)
            created = updated = skipped = 0
            runtime_errors = []
            classes = {c.name.strip().lower(): c for c in ClassRoom.query.filter_by(school_id=school.id)}
            for row in staged["valid"]:
                try:
                    username = row["username"].lower()
                    existing = User.query.filter_by(school_id=school.id, username=username).first()
                    admission_existing = (Student.query.filter_by(school_id=school.id, admission_no=row.get("admission_no")).first()
                                          if job.record_type == "student" else None)
                    duplicate = existing or admission_existing
                    if duplicate and mode == "skip":
                        skipped += 1
                        continue
                    password = row.get("password") or generate_temporary_password()
                    if job.record_type == "teacher":
                        account = existing
                        if account:
                            if account.role != "teacher":
                                raise ValueError("username belongs to another role")
                            account.full_name = row["full_name"]
                            account.email, account.phone = row.get("email", ""), row.get("phone", "")
                            updated += 1
                        else:
                            db.session.add(User(school_id=school.id, role="teacher", full_name=row["full_name"],
                                                username=username, email=row.get("email", ""), phone=row.get("phone", ""),
                                                password_hash=generate_password_hash(password), must_change_password=True))
                            created += 1
                    else:
                        student_record = admission_existing
                        account = db.session.get(User, student_record.user_id) if student_record else existing
                        if account and account.role != "student":
                            raise ValueError("username belongs to another role")
                        if student_record:
                            account.full_name, account.email, account.phone = row["full_name"], row.get("email", ""), row.get("phone", "")
                            student_record.class_id = classes[row["class"].lower()].id
                            student_record.guardian_name = row.get("guardian_name", "")
                            student_record.guardian_email = row.get("guardian_email", "")
                            student_record.guardian_phone = row.get("guardian_phone", "")
                            updated += 1
                        elif existing:
                            raise ValueError("student username exists with another admission number")
                        else:
                            account = User(school_id=school.id, role="student", full_name=row["full_name"],
                                           username=username, email=row.get("email", ""), phone=row.get("phone", ""),
                                           password_hash=generate_password_hash(password), must_change_password=True)
                            db.session.add(account)
                            db.session.flush()
                            db.session.add(Student(school_id=school.id, user_id=account.id,
                                                   class_id=classes[row["class"].lower()].id,
                                                   admission_no=row["admission_no"],
                                                   guardian_name=row.get("guardian_name", ""),
                                                   guardian_email=row.get("guardian_email", ""),
                                                   guardian_phone=row.get("guardian_phone", "")))
                            created += 1
                    db.session.commit()
                except Exception as exc:
                    db.session.rollback()
                    runtime_errors.append({"row": row.get("_row"), "error": str(exc)})
            db.session.commit()
            report = {"created": created, "updated": updated, "skipped": skipped,
                      "validation_errors": len(staged["errors"]), "runtime_errors": runtime_errors}
            job = db.session.get(BulkImportJob, job.id)
            job.status, job.duplicate_mode = "completed", mode
            job.report_json, job.completed_at = json.dumps(report), datetime.utcnow()
            log_action("bulk_import_completed", f"{job.record_type}: {created} created, {updated} updated, {skipped} skipped")
            db.session.commit()
            flash("Import completed. Review the report below.", "success")
            return redirect(url_for("bulk_import_preview", job_id=job.id))
        report = json.loads(job.report_json or "{}")
        body = """<main class="wrap"><div class="layout">""" + ctx["SIDEBAR"] + """<section class="grid"><article class="card"><h2>Import preview: {{ job.filename }}</h2><p><b>{{ staged.valid|length }}</b> valid rows · <b>{{ staged.errors|length }}</b> rows with errors</p>{% if job.status == 'preview' %}<form method="post">{{ csrf() }}<label>Duplicate handling<select name="duplicate_mode"><option value="skip">Skip existing records</option><option value="update">Update existing records</option></select></label><button class="btn green" {% if not staged.valid %}disabled{% endif %}>Commit valid rows</button></form>{% endif %}</article><article class="card"><h3>Validation errors</h3><table><tr><th>Row</th><th>Problems</th></tr>{% for item in staged.errors %}<tr><td>{{ item.row }}</td><td>{{ item.errors|join(', ') }}</td></tr>{% else %}<tr><td colspan="2">No validation errors.</td></tr>{% endfor %}</table></article>{% if report %}<article class="card"><h3>Import report</h3><p>Created: {{ report.created or 0 }} · Updated: {{ report.updated or 0 }} · Skipped: {{ report.skipped or 0 }} · Validation errors: {{ report.validation_errors or 0 }}</p><table><tr><th>Row</th><th>Runtime error</th></tr>{% for item in report.runtime_errors or [] %}<tr><td>{{ item.row }}</td><td>{{ item.error }}</td></tr>{% else %}<tr><td colspan="2">No runtime errors.</td></tr>{% endfor %}</table></article>{% endif %}</section></div></main>"""
        return render_page(body, title="Import Preview", job=job, staged=staged, report=report)

    app.student_report_is_paid = student_payment
    app.feature_models = models
